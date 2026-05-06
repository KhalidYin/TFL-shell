"""XLSX catalog and usage guide generator for TFL shell outputs v2.1."""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from tflshell import config
from tflshell.models.catalog import TFLCatalog
from tflshell.models.enums import Section
from tflshell.utils.naming import make_filename

HEADER_FILL = PatternFill(
    start_color=config.HEADER_BG_HEX, end_color=config.HEADER_BG_HEX, fill_type="solid"
)
HEADER_FONT = Font(name=config.FONT_NAME, size=9, bold=True, color=config.HEADER_FG_HEX)
DATA_FONT = Font(name=config.FONT_NAME, size=9)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
TOP_ALIGN = Alignment(vertical="top", wrap_text=True)
HEADER_ALIGN = Alignment(vertical="center", horizontal="center", wrap_text=True)
TYPE_FILLS = {
    "Table": PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"),
    "Figure": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "Listing": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
}

MASTER_COLUMNS = [
    "TFL ID",
    "Display Label",
    "Title",
    "Type",
    "Section",
    "Shell Family",
    "Study Phase Scope",
    "Coverage Summary",
    "Population",
    "Applicability",
    "Dataset Source",
    "Program Reference",
    "Dictionary / Standard",
    "Placeholder Style",
    "Footnotes",
    "Remarks",
]
MASTER_WIDTHS = [14, 18, 42, 10, 10, 22, 28, 32, 28, 18, 20, 22, 28, 24, 12, 30]
SECTION_SHEETS = [
    ("14.1_Demographics", Section.SEC_14_1),
    ("14.2_Efficacy", Section.SEC_14_2),
    ("14.3_Safety", Section.SEC_14_3),
    ("14.4_Special", Section.SEC_14_4),
    ("16.2_Listings", Section.SEC_16_2),
]


class XlsxTocGenerator:
    """Generate the governance-oriented TFL catalog workbook."""

    def __init__(self, catalog: TFLCatalog, output_path: str | None = None):
        self.catalog = catalog
        self.output_path = output_path or os.path.join(
            config.DEFAULT_OUTPUT_DIR,
            make_filename("TFL_TOC", config.VERSION, ".xlsx"),
        )

    def generate(self) -> str:
        wb = Workbook()

        toc_ws = wb.active
        toc_ws.title = "TOC_Master"
        self._build_master_sheet(toc_ws)

        for title, section in SECTION_SHEETS:
            ws = wb.create_sheet(title=title)
            self._build_catalog_sheet(ws, self.catalog.by_section(section))

        field_ws = wb.create_sheet(title="Field_Definitions")
        self._build_field_definitions_sheet(field_ws)

        usage_ws = wb.create_sheet(title="Usage_Guide")
        self._build_usage_guide_sheet(usage_ws)

        change_ws = wb.create_sheet(title="Change_Log")
        self._build_change_log_sheet(change_ws)

        os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
        wb.save(self.output_path)
        return self.output_path

    def _build_master_sheet(self, ws):
        self._build_catalog_sheet(ws, self.catalog.all())

    def _build_catalog_sheet(self, ws, items):
        self._write_headers(ws, MASTER_COLUMNS, MASTER_WIDTHS)

        for row_idx, item in enumerate(items, 2):
            values = self._catalog_row(item)
            row_fill = TYPE_FILLS.get(item.tfl_type.value)
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = DATA_FONT
                cell.alignment = TOP_ALIGN
                cell.border = THIN_BORDER
                if row_fill:
                    cell.fill = row_fill

        self._finalize_sheet(ws, len(items) + 1, len(MASTER_COLUMNS))

    def _build_field_definitions_sheet(self, ws):
        columns = ["Field", "Definition", "Example"]
        widths = [24, 72, 36]
        self._write_headers(ws, columns, widths)

        rows = [
            [
                "TFL ID",
                "Unique controlled identifier using type + CSR section + sequence.",
                "T14.3.1",
            ],
            [
                "Display Label",
                "Reviewer-facing shell label used in generated outputs.",
                "Table 14.3.1",
            ],
            [
                "Title",
                "Governance-approved descriptive shell title.",
                "Summary of Treatment-Emergent Adverse Events",
            ],
            ["Type", "Output class.", "Table / Figure / Listing"],
            ["Section", "CSR section aligned to ICH E3 scope.", "14.3"],
            [
                "Shell Family",
                "Governed shell-family grouping used for coverage and maintenance planning.",
                "Safety",
            ],
            [
                "Study Phase Scope",
                "High-level phase scope supported by the shell family or item.",
                "Phase I-III",
            ],
            [
                "Coverage Summary",
                "Governance summary derived from the approved coverage matrix.",
                "Core",
            ],
            ["Population", "Analysis set named in the shell header.", "Safety Population"],
            [
                "Applicability",
                "Whether the shell is general, oncology only, or non-oncology only.",
                "Oncology only",
            ],
            ["Dataset Source", "Primary ADaM/SDTM source or derived dataset lineage.", "ADAE"],
            [
                "Program Reference",
                "Planned derivation or output program reference.",
                "t_ae_summary.sas",
            ],
            [
                "Dictionary / Standard",
                "Coding dictionary or standard version documented for the shell.",
                "MedDRA 27.0; CTCAE 5.0",
            ],
            [
                "Placeholder Style",
                "Shell convention for non-first-column content.",
                "First column preserved; non-structural cells use style-appropriate shell placeholders such as XX or xx (xx.x)",
            ],
            ["Footnotes", "Whether shell-specific footnotes are present.", "Yes"],
            [
                "Remarks",
                "Additional implementation or governance notes.",
                "Simulated figure retained; shell family or study-specific notes if needed",
            ],
        ]
        self._write_rows(ws, rows)
        self._finalize_sheet(ws, len(rows) + 1, len(columns))

    def _build_usage_guide_sheet(self, ws):
        columns = ["Topic", "Guidance"]
        widths = [28, 96]
        self._write_headers(ws, columns, widths)

        rows = [
            [
                "Workbook purpose",
                "This workbook is a synchronized catalog and usage guide for the DOCX TFL shell template. It is not a workflow tracker.",
            ],
            [
                "Scope",
                "The workbook covers CSR Sections 14.1, 14.2, 14.3, 14.4, and 16.2 only; Section 16.1 is out of scope for this generator.",
            ],
            [
                "Coverage metadata",
                "Use Shell Family, Study Phase Scope, and Coverage Summary as governance metadata to interpret where a shell sits in the approved coverage matrix.",
            ],
            [
                "Applicability",
                "Use the Applicability column to identify shells relevant to oncology, non-oncology, or both. These labels guide study-specific shell selection rather than serving as workflow status flags.",
            ],
            [
                "Placeholder convention",
                "Tables and listings preserve structural examples in the first column and use adaptive shell placeholders in non-structural cells. Controlled treatment/group patterns use Group 1, Group 2, and may retain a separate ellipsis (...) expansion column; that expansion column must remain distinct and must not be merged with Overall, Total, HR, or other analytic columns. Header sample sizes remain generic rather than concrete.",
            ],
            [
                "Figures",
                "Figure entries correspond to simulated shell illustrations that should be replaced with study-specific outputs in production.",
            ],
            [
                "Ordering",
                "Row order matches the catalog ordering and should stay aligned with the generated DOCX template.",
            ],
            [
                "Word TOC",
                "After opening the DOCX outputs in Word, update fields so the automatic Table of Contents populates.",
            ],
            [
                "Change management",
                "Record template-level changes on the Change_Log sheet using one row per release decision or content update.",
            ],
        ]
        self._write_rows(ws, rows)
        self._finalize_sheet(ws, len(rows) + 1, len(columns))

    def _build_change_log_sheet(self, ws):
        columns = ["Date", "Version", "Scope", "Change Description", "Author"]
        widths = [16, 12, 18, 72, 24]
        self._write_headers(ws, columns, widths)

        today = datetime.now().strftime("%d %B %Y")
        rows = [
            [
                today,
                config.VERSION,
                "Master shell outputs",
                "Initial governance-aligned release covering Sections 14.1, 14.2, 14.3, 14.4, and 16.2 with shell-first adaptive placeholder semantics.",
                "TFLshell Generator",
            ]
        ]
        self._write_rows(ws, rows)
        self._finalize_sheet(ws, len(rows) + 1, len(columns))

    def _catalog_row(self, item):
        dict_str = (
            "; ".join(f"{k} {v}" for k, v in item.dictionary_versions.items())
            if item.dictionary_versions
            else ""
        )
        remarks = []
        if item.is_figure_generated:
            remarks.append("Simulated figure generated")
        if item.table_notes:
            remarks.append(item.table_notes)
        return [
            item.id,
            item.display_label,
            item.title,
            item.tfl_type.value,
            item.section.number,
            item.shell_family_label,
            item.study_phase_scope_label,
            item.coverage_summary_label,
            item.population,
            item.applicability_label,
            item.dataset_source,
            item.program_ref,
            dict_str,
            item.placeholder_summary,
            "Yes" if item.footnote_text() else "No",
            "; ".join(remarks),
        ]

    def _write_headers(self, ws, columns, widths):
        for col_idx, (name, width) in enumerate(zip(columns, widths), 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    def _write_rows(self, ws, rows):
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = DATA_FONT
                cell.alignment = TOP_ALIGN
                cell.border = THIN_BORDER

    def _finalize_sheet(self, ws, total_rows, total_cols):
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(total_cols)}{max(total_rows, 1)}"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
