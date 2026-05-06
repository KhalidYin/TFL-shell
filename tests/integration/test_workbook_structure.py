from openpyxl import load_workbook

from tflshell.data.definitions import build_catalog
from tflshell.generators.xlsx_toc import MASTER_COLUMNS, SECTION_SHEETS, XlsxTocGenerator


def test_workbook_contains_expected_sheets_and_headers(tmp_path):
    output_path = tmp_path / "tfl_toc.xlsx"
    generator = XlsxTocGenerator(build_catalog(), output_path=str(output_path))

    generated = generator.generate()
    workbook = load_workbook(generated, read_only=True, data_only=True)

    expected_sheets = [
        "TOC_Master",
        *[name for name, _ in SECTION_SHEETS],
        "Field_Definitions",
        "Usage_Guide",
        "Change_Log",
    ]
    assert workbook.sheetnames == expected_sheets

    master_sheet = workbook["TOC_Master"]
    header_row = [cell.value for cell in next(master_sheet.iter_rows(min_row=1, max_row=1))]
    assert header_row == MASTER_COLUMNS


def test_workbook_master_sheet_contains_governance_metadata(tmp_path):
    output_path = tmp_path / "tfl_toc.xlsx"
    generator = XlsxTocGenerator(build_catalog(), output_path=str(output_path))

    generated = generator.generate()
    workbook = load_workbook(generated, read_only=True, data_only=True)
    master_sheet = workbook["TOC_Master"]

    first_data_row = [cell.value for cell in next(master_sheet.iter_rows(min_row=2, max_row=2))]
    header_to_value = dict(zip(MASTER_COLUMNS, first_data_row))

    assert header_to_value["Shell Family"]
    assert header_to_value["Study Phase Scope"]
    assert header_to_value["Coverage Summary"]


def test_workbook_includes_new_phase_i_and_non_oncology_rows(tmp_path):
    output_path = tmp_path / "tfl_toc.xlsx"
    generator = XlsxTocGenerator(build_catalog(), output_path=str(output_path))

    generated = generator.generate()
    workbook = load_workbook(generated, read_only=True, data_only=True)
    master_sheet = workbook["TOC_Master"]

    rows = list(master_sheet.iter_rows(min_row=2, values_only=True))
    by_id = {row[0]: row for row in rows}

    assert by_id["T14.3.4.17"][5] == "Phase I Safety and Dose Escalation"
    assert by_id["T14.3.4.17"][6] == "Phase I"
    assert by_id["T14.2.23"][9] == "Non-Oncology only"
    assert by_id["T14.2.23"][7] == "Core (Phase II-III, Non-Oncology)"
